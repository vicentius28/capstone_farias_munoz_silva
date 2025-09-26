# views.py
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.views import APIView
from proyectoapp.models import ParticipacionCapacitacion
from proyectoapp.serializers import ParticipacionExportSerializer
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from openpyxl.styles import Alignment, Font

User = get_user_model()

class ExportarParticipacionesExcel(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    def get(self, request):
        titulo = request.GET.get("titulo")
        nombre = request.GET.get("nombre")
        anio = request.GET.get("anio")

        participaciones = ParticipacionCapacitacion.objects.select_related(
            "capacitacion__titulo_general"
        ).prefetch_related("usuario")

        if titulo:
            participaciones = participaciones.filter(
                capacitacion__titulo_general__titulo__iexact=titulo
            )

        if nombre:
            participaciones = participaciones.filter(
                capacitacion__nombre__iexact=nombre
            )

        if anio:
            participaciones = participaciones.filter(fecha_realizacion=anio)

        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participaciones"

        headers = ["Nombre Usuario", "Email"]
        ws.append(headers)

        total = 0
        for p in participaciones:
            usuarios = p.usuario.filter(empresa__id=2)
            for usuario in usuarios:
                ws.append([
                    f"{usuario.first_name} {usuario.last_name}",
                    usuario.email,
                ])
                total += 1

        ws.append([])
        ws.append(["Total usuarios:", total])

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 30

        # Nombre del archivo dinámico
        def slugify(value):
            return value.lower().replace(" ", "_").replace("/", "_")

        safe_nombre = slugify(nombre or "capacitacion")
        safe_anio = anio or "todos"
        filename = f"{safe_nombre}_{safe_anio}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class CapacitacionesAgrupadasAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    def get(self, request):
        participaciones = ParticipacionCapacitacion.objects.select_related(
            'capacitacion__titulo_general'
        ).all()

        data = {}

        for p in participaciones:
            titulo = p.capacitacion.titulo_general.titulo if p.capacitacion.titulo_general else "Sin título"
            nombre = p.capacitacion.nombre
            anio = p.fecha_realizacion

            # Agrupación por título
            if titulo not in data:
                data[titulo] = {}

            # Agrupación por nombre de capacitación
            if nombre not in data[titulo]:
                data[titulo][nombre] = set()

            # Agrega el año
            data[titulo][nombre].add(anio)

        # Formato final para el frontend
        response_data = []
        for titulo, capacitaciones in data.items():
            capacitaciones_list = []
            for nombre, anios in capacitaciones.items():
                capacitaciones_list.append({
                    "nombre": nombre,
                    "anios": [{"anio": a} for a in sorted(anios)]
                })
            response_data.append({
                "titulo": titulo,
                "capacitaciones": capacitaciones_list
            })

        return Response(response_data)
    
    
    
    
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
bold_font = Font(bold=True)

class ExportarResumenPorUsuarioAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        usuarios_activos = User.objects.filter(empresa__id=2, is_active=True).order_by("last_name", "first_name")
        participaciones = ParticipacionCapacitacion.objects.select_related(
            "capacitacion__titulo_general"
        ).prefetch_related("usuario")

        columnas_set = set()
        usuario_data = {}

        # Estructura: {(titulo, nombre)} y participaciones por usuario
        for p in participaciones:
            titulo = p.capacitacion.titulo_general.titulo if p.capacitacion.titulo_general else "Sin título"
            nombre = p.capacitacion.nombre
            key_col = (titulo, nombre)
            columnas_set.add(key_col)

            for u in p.usuario.filter(empresa__id=2, is_active=True):
                user_key = f"{u.first_name} {u.last_name}".strip().lower()
                if user_key not in usuario_data:
                    usuario_data[user_key] = {"usuario": u, "capacitaciones": {}}
                usuario_data[user_key]["capacitaciones"][key_col] = p.fecha_realizacion

        columnas = sorted(list(columnas_set))

        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen por Usuario"

        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold_font = Font(bold=True)

        # Encabezado: fila 1 (título) y fila 2 (nombre capacitación)
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 24

        ws.cell(row=1, column=1, value="Nombre Usuario")
        ws.cell(row=1, column=1).alignment = center_alignment
        ws.cell(row=1, column=1).font = bold_font

        for col_index, (titulo, _) in enumerate(columnas, start=2):
            cell = ws.cell(row=1, column=col_index, value=titulo)
            cell.alignment = center_alignment
            cell.font = bold_font

        for col_index, (_, nombre) in enumerate(columnas, start=2):
            cell = ws.cell(row=2, column=col_index, value=nombre)
            cell.alignment = center_alignment
            cell.font = bold_font

        # Cuerpo: todos los usuarios activos, con o sin participación
        for row_index, user in enumerate(usuarios_activos, start=3):
            user_key = f"{user.first_name} {user.last_name}".strip().lower()
            ws.cell(row=row_index, column=1, value=f"{user.first_name} {user.last_name}")
            ws.cell(row=row_index, column=1).alignment = center_alignment

            for col_index, col_key in enumerate(columnas, start=2):
                anio = usuario_data.get(user_key, {}).get("capacitaciones", {}).get(col_key)
                if anio:
                    cell = ws.cell(row=row_index, column=col_index, value=anio)
                    cell.alignment = center_alignment

        # Totales por capacitación (última fila)
        total_row = usuarios_activos.count() + 3
        ws.cell(row=total_row, column=1, value="Total:")
        ws.cell(row=total_row, column=1).alignment = center_alignment
        ws.cell(row=total_row, column=1).font = bold_font

        for col_index, col_key in enumerate(columnas, start=2):
            count = sum(
                1 for user in usuarios_activos
                if usuario_data.get(f"{user.first_name} {user.last_name}".strip().lower(), {})
                .get("capacitaciones", {}).get(col_key)
            )
            cell = ws.cell(row=total_row, column=col_index, value=count)
            cell.alignment = center_alignment
            cell.font = bold_font

        # Ajuste columnas
        for i in range(1, len(columnas) + 2):
            ws.column_dimensions[get_column_letter(i)].width = 30

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="resumen_usuarios.xlsx"'
        wb.save(response)
        return response
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication

from proyectoapp.models import ParticipacionCapacitacion

class ParticipantesAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        titulo = request.GET.get("titulo")
        nombre = request.GET.get("nombre")
        anio = request.GET.get("anio")
        
        if not (titulo and nombre and anio ):
            return Response({"usuarios": []})

        participaciones = ParticipacionCapacitacion.objects.filter(
            capacitacion__titulo_general__titulo__iexact=titulo,
            capacitacion__nombre__iexact=nombre,
            fecha_realizacion=anio
        ).prefetch_related("usuario")

        usuarios_list = []

        for participacion in participaciones:
            for usuario in participacion.usuario.filter(empresa__id=2):
                usuarios_list.append({
                    "nombre": f"{usuario.first_name} {usuario.last_name}".strip(),
                    "cargo": usuario.cargo.cargo
                })

        return Response({"usuarios": usuarios_list})
